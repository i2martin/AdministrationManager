from datetime import datetime
from io import BytesIO
import os
import PIL
from PIL.Image import Image
from openpyxl import load_workbook
from flask import Flask, render_template, request, redirect, url_for, send_file
from flask_wtf import FlaskForm
from flask import flash
from requests import get
from wtforms import StringField, SubmitField, SelectField, widgets, SelectMultipleField, PasswordField
import workdays as wd
import honorarium_excel as he
import travel_excel as te
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from flask_login import UserMixin, login_user, LoginManager, login_required, current_user, logout_user
import qrcode
import validators as validate
login_manager = LoginManager()
app = Flask(__name__)

logo_path = os.environ.get("LOGO_PATH2")


app.config['SECRET_KEY'] = 'C2HWGVoMGfNTBsrYQg8EcMrdTimkZfAb'

app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URL", 'sqlite:///users.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(100))
    name = db.Column(db.String(100))
    surname = db.Column(db.String(100))
    organisation = db.Column(db.String(100))
    work_address = db.Column(db.String(100))
    home_address = db.Column(db.String(100))
    travel_distance = db.Column(db.Float)
    transportation_fee = db.Column(db.Float)


class Inventory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    inventory_number = db.Column(db.String(100))
    name = db.Column(db.String(100))
    unit = db.Column(db.String(100))
    amount = db.Column(db.String(100))
    value = db.Column(db.String(100))
    location = db.Column(db.String(100))
    organisation = db.Column(db.String(100))
    item_status = db.Column(db.Boolean)  # True if inventory check is active and item is checked


class InventoryForm(FlaskForm):
    inventory_number = StringField('inventory_number')
    name = StringField('name')
    unit = StringField('unit')
    amount = StringField('amount')
    value = StringField('value')
    location = StringField('location')
    submit = SubmitField(label='Dodaj', id='submit')


class InventoryCheckHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    organisation = db.Column(db.String(100))
    inventory_check = db.Column(db.Boolean)  # True --> inventory check active
    user_initiated_check = db.Column(db.String(100))
    check_started_date = db.Column(db.String(100))
    check_finished_date = db.Column(db.String(100))


class InventoryCheckForm(FlaskForm):
    start_check = SubmitField(label='Započni provjeru', id='start_check')
    end_check = SubmitField(label='Završi provjeru', id='finish_check')
    generate_qrs = SubmitField(label='Dohvati QR kodove', id='generate_qrs')
    print_inventory = SubmitField(label='Ispiši inventar', id='print_inventory')


with app.app_context():
    db.create_all()
    login_manager.init_app(app)


class MultiCheckboxField(SelectMultipleField):
    widget = widgets.ListWidget(prefix_label=False)
    option_widget = widgets.CheckboxInput()


# TODO - add validators
class PutniTroskovi(FlaskForm):
    km_arrival = StringField('arrival')
    km_return = StringField('return')
    vehicle = SelectField('vehicle', choices=['Osobni automobil', 'Autobus', 'Vlak'])
    submit = SubmitField(label="Preuzmi", id='submit')


class LoginForm(FlaskForm):
    username = StringField('username')
    password = PasswordField('password')


class RegisterForm(FlaskForm):
    username = StringField('username')
    organisation = StringField('username')
    password = PasswordField('password')
    confirm_password = PasswordField('confirm_password')


class Postavke(FlaskForm):
    name = StringField('arrival')
    surname = StringField('arrival')
    work_address = StringField('arrival')
    home_address = StringField('arrival')
    travel_distance = StringField('arrival')
    submit = SubmitField(label='Spremi', id='submit')
    transportation_fee = StringField('transportation_fee')


# TODO - add validators
class Honorari(FlaskForm):
    subject = StringField('subject')
    class_tag = StringField('class-tag')
    hours = StringField('date')
    submit = SubmitField(label='Preuzmi', id='submit')


@app.route('/')
def home():
    return render_template('index.html', logo = os.environ.get('LOGO_PATH'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm()
    # TODO: Add checks (i.e. existing username) so that it doesn't crash
    if request.method == "POST" and form.validate() and validate.check_username(request.form.get('username'))\
            and validate.check_password(request.form.get('password')):
        new_user = User(
            username=request.form.get('username'),
            password=generate_password_hash(password=request.form.get('password'), method='pbkdf2:sha256',
                                            salt_length=8),
            organisation=request.form.get('organisation')
        )
        user = User.query.filter_by(username=new_user.username).first()
        # check if username exists
        if user:
            flash('Korisničko ime već postoji. Odaberite drugo korisničko ime!')
            return redirect(url_for('register'))
        # username exists and password is correct
        else:
            db.session.add(new_user)
            db.session.commit()
            return redirect(url_for('login'))
    else:
        # this means either it was GET request or checks failed
        if request.method == "GET":
            return render_template('register.html',logo_path=logo_path, form=form)
        else:
            flash('Dogodila se greška prilikom registracije.')
            return redirect(url_for('register'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST' and validate.check_password(request.form.get('password')) and validate.check_username(request.form.get('username')):
        username = request.form.get('username')
        password = request.form.get('password')
        # Find user by username entered.
        user = User.query.filter_by(username=username).first()
        # check if username and password are correct
        if not user or not check_password_hash(user.password, password):
            flash('Neispravno korisničko ime ili lozinka. Molimo pokušajte ponovno!')
            return redirect(url_for('login'))
        # username exists and password is correct
        else:
            login_user(user)
            return redirect(url_for('view_services'))
    return render_template('login.html', logo_path = logo_path)


@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('home'))


@app.route('/travel_expense', methods=['GET', 'POST'])
@login_required
def travel_expense():
    workdays = []
    year = datetime.now().year
    month = datetime.now().month
    for d in wd.get_workdays(year, month, ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']):
        workdays.append(d.strftime('%d.%m.%Y'))
    workdays = wd.order_days(workdays)
    number_of_workdays = len(workdays)
    if request.method == 'POST':
        return te.travel(data=request.form.to_dict(flat=False), workdays=workdays, user=current_user)
    else:
        form = PutniTroskovi()
        return render_template('travel.html',logo_path = logo_path, form=form, workdays=workdays, number_of_workdays=number_of_workdays,
                               travel_distance=current_user.travel_distance)


@app.route('/honorarium', methods=['GET', 'POST'])
@login_required
def honorarium():
    workdays = []
    year = datetime.now().year
    month = datetime.now().month
    for d in wd.get_workdays(year, month, ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']):
        workdays.append(d.strftime('%d.%m.%Y'))
    workdays = wd.order_days(workdays)
    workdays = wd.only_date(workdays)
    number_of_workdays = len(workdays)
    if request.method == 'POST':
        return he.honorarium(request.form.to_dict(flat=False), workdays, user=current_user)
        #return send_from_directory(directory='static', path='files/temp_files/tablica-honorari.xlsx')
        #return redirect((url_for('honorari')))
    else:
        forms = []
        for i in range(0, 5):
            form = Honorari()
            forms.append(form)
        return render_template('honorarium.html', logo_path = logo_path, forms=forms, workdays=workdays, number_of_workdays=number_of_workdays)


@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    form = Postavke()
    if request.method == 'POST':
        name = request.form.get('name')
        surname = request.form.get('surname')
        work_address = request.form.get('work_address')
        home_address = request.form.get('home_address')
        travel_distance = request.form.get('travel_distance')
        transportation_fee = request.form.get('transportation_fee')
        if name != "":
            current_user.name = name
            db.session.commit()
        if surname != "":
            current_user.surname = surname
            db.session.commit()
        if work_address != "":
            current_user.work_address = work_address
            db.session.commit()
        if home_address != "":
            current_user.home_address = home_address
            db.session.commit()
        if travel_distance != "":
            current_user.travel_distance = travel_distance
            db.session.commit()
        if transportation_fee != "":
            current_user.transportation_fee = transportation_fee
            db.session.commit()
        return redirect(url_for('view_services'))
    return render_template('settings.html', form=form, logo_path = logo_path)


@app.route('/services', methods=['GET'])
@login_required
def view_services():
    return render_template('services.html', logo_path = logo_path)


@app.route('/add_inventory', methods=["GET", "POST"])
@login_required
def add_inventory():
    form = InventoryForm()
    if request.method == 'POST':
        if Inventory.query.filter_by(inventory_number=request.form.get('inventory_number'), organisation=current_user.organisation).first():
            flash('Inventarni broj već postoji!')
            return redirect(url_for('add_inventory'))
        else:
            new_inventory = Inventory(
                inventory_number=request.form.get('inventory_number'),
                name=request.form.get('name'),
                unit=request.form.get('unit'),
                amount=request.form.get('amount'),
                value=request.form.get('value'),
                location=request.form.get('location'),
                organisation=current_user.organisation,
                item_status=False)
            db.session.add(new_inventory)
            db.session.commit()
            flash('Inventar je uspješno ažuriran!')
        return redirect(url_for('add_inventory', status="success"))
    return render_template('inventory.html', form=form, logo_path = logo_path)


@app.route("/view_inventory")
@login_required
def view_inventory():
    form = InventoryCheckForm()
    form_popup = InventoryForm()
    if InventoryCheckHistory.query.filter_by(organisation=current_user.organisation).first():
        status = InventoryCheckHistory.query.filter_by(organisation=current_user.organisation).first().inventory_check
    else:
        status = False
    # get complete inventory for users organisation
    inventory = Inventory.query.filter_by(organisation=current_user.organisation).all()
    # get list of specific locations
    locations = []
    inventory_by_location = []
    for item in inventory:
        if item.location not in locations:
            inventory_list = []
            locations.append(item.location)
            # find all inventory by specific organisation and location
            for i in Inventory.query.filter_by(organisation=current_user.organisation, location=item.location).all():
                inventory_list.append(i)
            inventory_by_location.append(inventory_list)
    return render_template('viewInventory.html',logo_path = logo_path, organisation=current_user.organisation, locations=locations,
                           data=inventory_by_location, status=status, form=form, form_popup=form_popup)


@app.route('/inventory/<organisation>/<inventory_id>')
def inventory_check(organisation, inventory_id):
    item = Inventory.query.filter_by(inventory_number=inventory_id, organisation=organisation).first()
    if item:
        if InventoryCheckHistory.query.filter_by(organisation=organisation).first().inventory_check is True:
        # update item status to checked (True)
            item.item_status = True
            db.session.commit()
        return render_template('itemStatus.html', name=item.name)
    else:
        return render_template('itemStatus.html')


@app.route('/inventory_check_status')
@login_required
def inventory_check_status():
    record = InventoryCheckHistory.query.filter_by(organisation=current_user.organisation).first()
    if record is None or record == 'None' or record == 'NoneType':
        new_check = InventoryCheckHistory(
            organisation=current_user.organisation,
            inventory_check=True,
            user_initiated_check=current_user.username,
            check_started_date=datetime.utcnow(),
            check_finished_date="")
        db.session.add(new_check)
        db.session.commit()
    elif record.inventory_check is not True:
        record.inventory_check = True
        record.check_started_date = datetime.utcnow()
        record.check_finished_date = ""
        db.session.commit()
    else:
        record.inventory_check = False
        record.check_finished_date = datetime.utcnow()
        db.session.commit()
        # TODO: print inventory check report --> missing/unchecked items (maybe ask to finish it with missing items
        #  or not?)
        # TODO: find a better way to link to a file
        request = get('https://administration-manager.herokuapp.com/static/files/inventar.xlsx')
        buffer = BytesIO(request.content)
        workbook = load_workbook(buffer)
        worksheet = workbook['Inventar']
        items = Inventory.query.filter_by(organisation=current_user.organisation).all()
        i = 1
        for item in items:
            i = i + 1
            worksheet["A" + str(i)] = item.inventory_number
            worksheet["B" + str(i)] = item.location
            worksheet["C" + str(i)] = item.name
            worksheet["D" + str(i)] = item.unit
            worksheet["E" + str(i)] = item.amount
            worksheet["F" + str(i)] = item.value
            if item.item_status == True:
                worksheet["G" + str(i)] = "✔"
            else:
                worksheet["G" + str(i)] = "❌"
        buffer2 = BytesIO()
        workbook.save(buffer2)
        buffer2.seek(0)
        # loop through all the inventory items and change their status to unchecked --> False
        for item in items:
            item.item_status = False
            db.session.commit()
        print(request)
        return send_file(BytesIO(buffer2.read()), mimetype="application/vnd.ms-excel", download_name="inventar.xlsx", as_attachment=False)
    return redirect('view_inventory')


def concat_images_horizontally(images):
    """
    Function that accepts a list of PIL.Image objects and stacks them up horizontally in rows of 2 images.
    :param images: list of PIL.Image objects
    :return: list of stacked up PIL.Image objects
    """
    new_images = []
    if len(images) >= 2:
        for i in range(0, len(images)):
            global dst
            if i % 2 == 0 or i == len(images) - 1:
                if i != len(images) - 1:
                    dst = PIL.Image.new('RGB', (160, 80))
                    dst.paste(images[i], (0, 0))
                else:
                    dst = PIL.Image.new('RGB', (80, 80))
                    dst.paste(images[i], (0, 0))
                    new_images.append(dst)
            else:
                dst.paste(images[i], (80,0))
                new_images.append(dst)
        return new_images
    else:
        return images


def concat_images_vertically(images):
    """
    Function that accepts a list of PIL.Image objects and stacks them up vertically in columns of 2 images.
    :param images: list of PIL.Image objects
    :return: list of vertically stacked PIL.Image objects
    """
    global dst
    new_images = []
    if len(images) >= 2:
        for i in range(0, len(images)):
            if i % 2 == 0 or i == len(images) - 1:
                if i != len(images) - 1:
                    dst = PIL.Image.new('RGB', (160, 160))
                    dst.paste(images[i], (0, 0))
                elif len(images) - 1 == 1:
                    dst.paste(images[i], (0, 80))
                    dst.paste(PIL.Image.new("RGB", (80, 80), (0, 0, 0)), (80, 80))
                    new_images.append(dst)
                else:
                    dst = PIL.Image.new('RGB', (160, 160))
                    dst.paste(images[i], (0, 0))
                    new_images.append(dst)
            else:
                dst.paste(images[i], (0, 80))
                new_images.append(dst)
        return new_images
    else:
        return images


@app.route('/generate_qr_codes')
@login_required
def generate_qr_codes():
    data = Inventory.query.filter_by(organisation=current_user.organisation).all()
    qrcodes = []
    for i in range(0, len(data)):
        img = qrcode.make("https://administration-manager.herokuapp.com/inventory/" + current_user.organisation + "/" + data[i].inventory_number)
        img = img.resize(size=(80, 80))  # TODO: this might cause a problem with size
        qrcodes.append(img)
    if len(qrcodes) > 0:
        if len(qrcodes) > 2:
            qrcodes = concat_images_horizontally(qrcodes)
            if len(qrcodes) >= 2:
                qrcodes = concat_images_vertically(qrcodes)
        qrcodes_buffer = BytesIO()
        qrcodes = qrcodes
        qrcodes[0].save(qrcodes_buffer, "PDF", resolution=100.0, save_all=True, append_images=qrcodes[0:])
        qrcodes_buffer.seek(0)
        response = send_file(qrcodes_buffer, mimetype='application/pdf')
        return response
    else:
        return redirect('view_inventory')


@app.route('/print_inventory')
@login_required
def print_inventory():
    data = Inventory.query.filter_by(organisation=current_user.organisation).order_by(Inventory.location.asc()).all()
    # TODO: find a better way to link to a file
    request = get('https://administration-manager.herokuapp.com/static/files/inventar.xlsx')
    buffer = BytesIO(request.content)
    workbook = load_workbook(buffer)
    worksheet = workbook['Inventar']
    i = 1
    for item in data:
        i = i + 1
        worksheet["A" + str(i)] = item.inventory_number
        worksheet["B" + str(i)] = item.location
        worksheet["C" + str(i)] = item.name
        worksheet["D" + str(i)] = item.unit
        worksheet["E" + str(i)] = item.amount
        worksheet["F" + str(i)] = item.value
    buffer2 = BytesIO()
    workbook.save(buffer2)
    buffer2.seek(0)
    return send_file(BytesIO(buffer2.read()), mimetype="application/vnd.ms-excel", download_name="inventar.xlsx", as_attachment=True)


@app.route('/remove_inventory/int:<id>')
@login_required
def remove_inventory(id):
    item = Inventory.query.filter_by(inventory_number = id, organisation = current_user.organisation).first()
    db.session.delete(item)
    db.session.commit()
    return redirect(url_for('view_inventory'))


@app.route('/update_inventory/<id>', methods=["POST"])
@login_required
# TODO: Finish this to redirect to a form that will update existing item with set changes
def update_inventory(id):
    if request.method == "POST":
        item = Inventory.query.filter_by(inventory_number = id, organisation=current_user.organisation).first()
        # TODO: Check if the new inventory number already exists
        item.inventory_number = request.form.get('inventory_number')
        item.name = request.form.get('name')
        item.location = request.form.get('location')
        item.amount = request.form.get('amount')
        item.unit = request.form.get('unit')
        item.value = request.form.get('value')
        db.session.commit()
        return redirect(url_for('view_inventory'))


if __name__ == '__main__':
    app.run()
